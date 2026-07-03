"""
Scanner de Arquivos da Rede de Tarifas - ARPE
==============================================
Detecta:
  1. Arquivos DUPLICADOS (mesmo conteúdo, nomes diferentes) via hash SHA-256
  2. Arquivos com MÚLTIPLAS VERSÕES (nomes similares: v1, v2, _copia, (1), etc.)
  3. Arquivos mais PESADOS da pasta

Uso:
  python scanner_rede.py "Z:\\Tarifas"
  python scanner_rede.py "Z:\\Tarifas" --top 20 --export
"""

import os
import sys
import hashlib
import re
import csv
from collections import defaultdict
from datetime import datetime


# ══════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════

TOP_PESADOS = 15          # quantos arquivos pesados mostrar
BLOCO_LEITURA = 65536     # 64KB por leitura (eficiente pra arquivos grandes)
MIN_TAMANHO_DUPLICATA = 1024  # ignora arquivos < 1KB na busca de duplicatas

# Padrões que indicam "versões" do mesmo arquivo
VERSAO_PATTERNS = [
    r'\s*\(\d+\)',              # " (1)", " (2)"
    r'\s*-\s*[Cc][oó]pia',     # " - Cópia", " - copia"
    r'\s*_v\d+',               # "_v1", "_v2"
    r'\s*v\d+',                # " v1", "v2"
    r'\s*_\d{8}',              # "_20240315" (data)
    r'\s*_old',                # "_old"
    r'\s*_novo',               # "_novo"
    r'\s*_final',              # "_final"
    r'\s*_revisado',           # "_revisado"
    r'\s*_corrigido',          # "_corrigido"
    r'\s*_atualizado',         # "_atualizado"
    r'\s*_backup',             # "_backup"
    r'\s*_bkp',                # "_bkp"
    r'\s*\s+FINAL',            # " FINAL"
    r'\s*\s+REVISADO',         # " REVISADO"
    r'\s*\s+OK',               # " OK"
    r'\s*\s+NOVO',             # " NOVO"
]


# ══════════════════════════════════════════════════════════════════
# CORES DO TERMINAL
# ══════════════════════════════════════════════════════════════════

class Cor:
    RESET   = '\033[0m'
    BOLD    = '\033[1m'
    RED     = '\033[91m'
    GREEN   = '\033[92m'
    YELLOW  = '\033[93m'
    CYAN    = '\033[96m'
    MAGENTA = '\033[95m'
    DIM     = '\033[2m'

    @staticmethod
    def suporta_cor():
        """Verifica se o terminal suporta cores."""
        if os.name == 'nt':
            os.system('')  # habilita ANSI no Windows 10+
        return hasattr(sys.stdout, 'isatty') and sys.stdout.isatty()


def c(texto, cor):
    """Aplica cor se suportado."""
    if Cor.suporta_cor():
        return f"{cor}{texto}{Cor.RESET}"
    return texto


# ══════════════════════════════════════════════════════════════════
# FUNÇÕES UTILITÁRIAS
# ══════════════════════════════════════════════════════════════════

def formatar_tamanho(bytes_val):
    """Converte bytes para formato legível."""
    for unidade in ['B', 'KB', 'MB', 'GB', 'TB']:
        if bytes_val < 1024:
            return f"{bytes_val:.1f} {unidade}"
        bytes_val /= 1024
    return f"{bytes_val:.1f} PB"


def calcular_hash(caminho):
    """Calcula SHA-256 do arquivo."""
    sha = hashlib.sha256()
    try:
        with open(caminho, 'rb') as f:
            while True:
                bloco = f.read(BLOCO_LEITURA)
                if not bloco:
                    break
                sha.update(bloco)
        return sha.hexdigest()
    except (PermissionError, OSError):
        return None


def normalizar_nome(nome_arquivo):
    """Remove sufixos de versão para agrupar arquivos similares."""
    nome, ext = os.path.splitext(nome_arquivo)
    for pattern in VERSAO_PATTERNS:
        nome = re.sub(pattern, '', nome, flags=re.IGNORECASE)
    return nome.strip().lower() + ext.lower()


def data_modificacao(caminho):
    """Retorna data de modificação formatada."""
    try:
        ts = os.path.getmtime(caminho)
        return datetime.fromtimestamp(ts).strftime('%d/%m/%Y %H:%M')
    except OSError:
        return 'N/A'


# ══════════════════════════════════════════════════════════════════
# SCANNER PRINCIPAL
# ══════════════════════════════════════════════════════════════════

def escanear_pasta(caminho_raiz):
    """Percorre toda a pasta e coleta informações dos arquivos."""
    arquivos = []
    erros = []
    total = 0

    print(f"\n{c('⏳ Escaneando...', Cor.CYAN)} {caminho_raiz}")
    print(c('   (isso pode demorar dependendo do tamanho da rede)\n', Cor.DIM))

    for raiz, dirs, nomes in os.walk(caminho_raiz):
        for nome in nomes:
            total += 1
            caminho_completo = os.path.join(raiz, nome)
            try:
                tamanho = os.path.getsize(caminho_completo)
                arquivos.append({
                    'nome': nome,
                    'caminho': caminho_completo,
                    'tamanho': tamanho,
                    'nome_normalizado': normalizar_nome(nome),
                    'pasta': raiz,
                })
            except (PermissionError, OSError) as e:
                erros.append((caminho_completo, str(e)))

            if total % 500 == 0:
                print(f"   {c(f'{total}', Cor.DIM)} arquivos lidos...", end='\r')

    print(f"   {c(f'{total}', Cor.GREEN)} arquivos encontrados.          \n")
    return arquivos, erros


def encontrar_duplicados_hash(arquivos):
    """Agrupa arquivos com conteúdo idêntico via SHA-256."""
    # Fase 1: agrupa por tamanho (pré-filtro rápido)
    por_tamanho = defaultdict(list)
    for arq in arquivos:
        if arq['tamanho'] >= MIN_TAMANHO_DUPLICATA:
            por_tamanho[arq['tamanho']].append(arq)

    candidatos = {tam: grupo for tam, grupo in por_tamanho.items() if len(grupo) > 1}

    # Fase 2: calcula hash só dos que têm tamanho igual
    total_hash = sum(len(g) for g in candidatos.values())
    print(f"   {c('Calculando hashes', Cor.CYAN)} de {total_hash} candidatos...")

    por_hash = defaultdict(list)
    processados = 0
    for grupo in candidatos.values():
        for arq in grupo:
            h = calcular_hash(arq['caminho'])
            if h:
                arq['hash'] = h
                por_hash[h].append(arq)
            processados += 1
            if processados % 100 == 0:
                print(f"   {processados}/{total_hash} hashes...", end='\r')

    duplicados = {h: grupo for h, grupo in por_hash.items() if len(grupo) > 1}
    print(f"   {c('Hashes concluídos.', Cor.GREEN)}                        \n")
    return duplicados


def encontrar_versoes(arquivos):
    """Agrupa arquivos cujo nome, sem sufixos de versão, é o mesmo."""
    por_nome = defaultdict(list)
    for arq in arquivos:
        chave = (arq['pasta'], arq['nome_normalizado'])
        por_nome[chave].append(arq)

    return {k: sorted(v, key=lambda x: x['nome'])
            for k, v in por_nome.items() if len(v) > 1}


def top_pesados(arquivos, n=TOP_PESADOS):
    """Retorna os N arquivos mais pesados."""
    return sorted(arquivos, key=lambda x: x['tamanho'], reverse=True)[:n]


# ══════════════════════════════════════════════════════════════════
# RELATÓRIO NO TERMINAL
# ══════════════════════════════════════════════════════════════════

def barra(titulo):
    largura = 70
    print(f"\n{'═' * largura}")
    print(f"  {c(titulo, Cor.BOLD + Cor.CYAN)}")
    print(f"{'═' * largura}")


def exibir_duplicados(duplicados):
    barra(f"🔴 ARQUIVOS DUPLICADOS (conteúdo idêntico) — {len(duplicados)} grupo(s)")

    if not duplicados:
        print(f"  {c('Nenhum duplicado encontrado!', Cor.GREEN)}")
        return

    espaco_desperdicado = 0

    for i, (h, grupo) in enumerate(duplicados.items(), 1):
        tam = grupo[0]['tamanho']
        desperdicio = tam * (len(grupo) - 1)
        espaco_desperdicado += desperdicio

        print(f"\n  {c(f'Grupo {i}', Cor.YELLOW)} — {len(grupo)} cópias — "
              f"cada uma: {c(formatar_tamanho(tam), Cor.MAGENTA)} — "
              f"desperdício: {c(formatar_tamanho(desperdicio), Cor.RED)}")

        for arq in grupo:
            mod = data_modificacao(arq['caminho'])
            print(f"    • {arq['nome']}")
            print(f"      {c(arq['caminho'], Cor.DIM)}")
            print(f"      Modificado: {mod}")

    print(f"\n  {c('TOTAL desperdiçado:', Cor.RED + Cor.BOLD)} "
          f"{c(formatar_tamanho(espaco_desperdicado), Cor.RED + Cor.BOLD)}")


def exibir_versoes(versoes):
    barra(f"🟡 ARQUIVOS COM MÚLTIPLAS VERSÕES — {len(versoes)} grupo(s)")

    if not versoes:
        print(f"  {c('Nenhum arquivo com versões múltiplas!', Cor.GREEN)}")
        return

    for i, ((pasta, nome_base), grupo) in enumerate(
            sorted(versoes.items(), key=lambda x: len(x[1]), reverse=True), 1):
        if i > 30:
            print(f"\n  {c(f'... e mais {len(versoes) - 30} grupos', Cor.DIM)}")
            break

        print(f"\n  {c(f'Grupo {i}', Cor.YELLOW)} — base: {c(nome_base, Cor.CYAN)} "
              f"— {len(grupo)} versões")
        print(f"  Pasta: {c(pasta, Cor.DIM)}")

        for arq in grupo:
            tam = formatar_tamanho(arq['tamanho'])
            mod = data_modificacao(arq['caminho'])
            print(f"    • {arq['nome']}  [{tam}]  mod: {mod}")


def exibir_pesados(pesados):
    barra(f"🔵 TOP {len(pesados)} ARQUIVOS MAIS PESADOS")

    for i, arq in enumerate(pesados, 1):
        tam = formatar_tamanho(arq['tamanho'])
        print(f"  {c(f'{i:>2}.', Cor.BOLD)} {c(tam, Cor.MAGENTA):>12}  {arq['nome']}")
        print(f"      {c(arq['caminho'], Cor.DIM)}")


def exibir_resumo(arquivos, duplicados, versoes, erros):
    barra("📊 RESUMO GERAL")

    total_tam = sum(a['tamanho'] for a in arquivos)
    desp = sum(g[0]['tamanho'] * (len(g) - 1) for g in duplicados.values())

    print(f"  Total de arquivos:        {c(str(len(arquivos)), Cor.BOLD)}")
    print(f"  Espaço total:             {c(formatar_tamanho(total_tam), Cor.BOLD)}")
    print(f"  Grupos duplicados:        {c(str(len(duplicados)), Cor.RED if duplicados else Cor.GREEN)}")
    print(f"  Espaço desperdiçado:      {c(formatar_tamanho(desp), Cor.RED if desp else Cor.GREEN)}")
    print(f"  Grupos com versões:       {c(str(len(versoes)), Cor.YELLOW if versoes else Cor.GREEN)}")
    if erros:
        print(f"  Arquivos inacessíveis:    {c(str(len(erros)), Cor.RED)}")
    print()


# ══════════════════════════════════════════════════════════════════
# EXPORTAÇÃO CSV
# ══════════════════════════════════════════════════════════════════

def exportar_excel(duplicados, versoes, pesados, arquivos, pasta_saida):
    """Gera Excel formatado com abas: Duplicados, Versões, Pesados, Resumo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    caminho_xlsx = os.path.join(pasta_saida, f'scanner_rede_{timestamp}.xlsx')

    wb = Workbook()

    # Estilos
    titulo_font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    titulo_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    erro_fill = PatternFill('solid', fgColor='FCE4EC')
    alerta_fill = PatternFill('solid', fgColor='FFF9C4')
    borda = Border(
        bottom=Side(style='thin', color='D9D9D9')
    )
    corpo_font = Font(name='Arial', size=10)

    def estilizar_header(ws, linha, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=linha, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def titulo_aba(ws, texto, num_cols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        cell = ws.cell(row=1, column=1, value=texto)
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

    # ── ABA 1: DUPLICADOS ──
    ws_dup = wb.active
    ws_dup.title = 'Duplicados'
    cols_dup = ['Grupo', 'Arquivo', 'Caminho Completo', 'Tamanho', 'Desperdício', 'Hash SHA-256']
    titulo_aba(ws_dup, f'🔴 ARQUIVOS DUPLICADOS — {len(duplicados)} grupo(s)', len(cols_dup))

    for col_idx, nome in enumerate(cols_dup, 1):
        ws_dup.cell(row=3, column=col_idx, value=nome)
    estilizar_header(ws_dup, 3, len(cols_dup))

    linha = 4
    cor_grupo = False
    for i, (h, grupo) in enumerate(duplicados.items(), 1):
        tam = grupo[0]['tamanho']
        desperdicio = formatar_tamanho(tam * (len(grupo) - 1))
        cor_grupo = not cor_grupo
        for j, arq in enumerate(grupo):
            ws_dup.cell(row=linha, column=1, value=i).font = corpo_font
            ws_dup.cell(row=linha, column=2, value=arq['nome']).font = corpo_font
            ws_dup.cell(row=linha, column=3, value=arq['caminho']).font = corpo_font
            ws_dup.cell(row=linha, column=4, value=formatar_tamanho(arq['tamanho'])).font = corpo_font
            ws_dup.cell(row=linha, column=5, value=desperdicio if j == 0 else '').font = corpo_font
            ws_dup.cell(row=linha, column=6, value=h[:16] + '...').font = Font(name='Arial', size=8, color='888888')
            if cor_grupo:
                for c_idx in range(1, len(cols_dup) + 1):
                    ws_dup.cell(row=linha, column=c_idx).fill = PatternFill('solid', fgColor='F2F7FB')
            for c_idx in range(1, len(cols_dup) + 1):
                ws_dup.cell(row=linha, column=c_idx).border = borda
            linha += 1

    ws_dup.column_dimensions['A'].width = 8
    ws_dup.column_dimensions['B'].width = 40
    ws_dup.column_dimensions['C'].width = 70
    ws_dup.column_dimensions['D'].width = 14
    ws_dup.column_dimensions['E'].width = 14
    ws_dup.column_dimensions['F'].width = 20
    ws_dup.auto_filter.ref = f'A3:F{max(linha - 1, 3)}'
    ws_dup.freeze_panes = 'A4'

    # ── ABA 2: VERSÕES ──
    ws_ver = wb.create_sheet('Versões')
    cols_ver = ['Grupo', 'Nome Base', 'Arquivo', 'Pasta', 'Tamanho', 'Modificado']
    titulo_aba(ws_ver, f'🟡 MÚLTIPLAS VERSÕES — {len(versoes)} grupo(s)', len(cols_ver))

    for col_idx, nome in enumerate(cols_ver, 1):
        ws_ver.cell(row=3, column=col_idx, value=nome)
    estilizar_header(ws_ver, 3, len(cols_ver))

    linha = 4
    cor_grupo = False
    for i, ((pasta, nome_base), grupo) in enumerate(
            sorted(versoes.items(), key=lambda x: len(x[1]), reverse=True), 1):
        cor_grupo = not cor_grupo
        for arq in grupo:
            ws_ver.cell(row=linha, column=1, value=i).font = corpo_font
            ws_ver.cell(row=linha, column=2, value=nome_base).font = corpo_font
            ws_ver.cell(row=linha, column=3, value=arq['nome']).font = corpo_font
            ws_ver.cell(row=linha, column=4, value=arq['pasta']).font = corpo_font
            ws_ver.cell(row=linha, column=5, value=formatar_tamanho(arq['tamanho'])).font = corpo_font
            ws_ver.cell(row=linha, column=6, value=data_modificacao(arq['caminho'])).font = corpo_font
            if cor_grupo:
                for c_idx in range(1, len(cols_ver) + 1):
                    ws_ver.cell(row=linha, column=c_idx).fill = PatternFill('solid', fgColor='FFFDE7')
            for c_idx in range(1, len(cols_ver) + 1):
                ws_ver.cell(row=linha, column=c_idx).border = borda
            linha += 1

    ws_ver.column_dimensions['A'].width = 8
    ws_ver.column_dimensions['B'].width = 35
    ws_ver.column_dimensions['C'].width = 45
    ws_ver.column_dimensions['D'].width = 60
    ws_ver.column_dimensions['E'].width = 14
    ws_ver.column_dimensions['F'].width = 18
    ws_ver.auto_filter.ref = f'A3:F{max(linha - 1, 3)}'
    ws_ver.freeze_panes = 'A4'

    # ── ABA 3: PESADOS ──
    ws_pes = wb.create_sheet('Pesados')
    cols_pes = ['Rank', 'Arquivo', 'Caminho Completo', 'Tamanho', 'Extensão']
    titulo_aba(ws_pes, f'🔵 TOP {len(pesados)} ARQUIVOS MAIS PESADOS', len(cols_pes))

    for col_idx, nome in enumerate(cols_pes, 1):
        ws_pes.cell(row=3, column=col_idx, value=nome)
    estilizar_header(ws_pes, 3, len(cols_pes))

    linha = 4
    for i, arq in enumerate(pesados, 1):
        _, ext = os.path.splitext(arq['nome'])
        ws_pes.cell(row=linha, column=1, value=i).font = corpo_font
        ws_pes.cell(row=linha, column=2, value=arq['nome']).font = corpo_font
        ws_pes.cell(row=linha, column=3, value=arq['caminho']).font = corpo_font
        ws_pes.cell(row=linha, column=4, value=formatar_tamanho(arq['tamanho'])).font = corpo_font
        ws_pes.cell(row=linha, column=5, value=ext.lower()).font = corpo_font
        if i <= 3:
            for c_idx in range(1, len(cols_pes) + 1):
                ws_pes.cell(row=linha, column=c_idx).fill = erro_fill
        for c_idx in range(1, len(cols_pes) + 1):
            ws_pes.cell(row=linha, column=c_idx).border = borda
        linha += 1

    ws_pes.column_dimensions['A'].width = 8
    ws_pes.column_dimensions['B'].width = 45
    ws_pes.column_dimensions['C'].width = 70
    ws_pes.column_dimensions['D'].width = 14
    ws_pes.column_dimensions['E'].width = 12
    ws_pes.freeze_panes = 'A4'

    # ── ABA 4: RESUMO ──
    ws_res = wb.create_sheet('Resumo')
    titulo_aba(ws_res, '📊 RESUMO DA VARREDURA', 3)

    total_tam = sum(a['tamanho'] for a in arquivos)
    desp = sum(g[0]['tamanho'] * (len(g) - 1) for g in duplicados.values())

    dados_resumo = [
        ('Total de arquivos', len(arquivos)),
        ('Espaço total', formatar_tamanho(total_tam)),
        ('Grupos duplicados', len(duplicados)),
        ('Espaço desperdiçado', formatar_tamanho(desp)),
        ('Grupos com versões', len(versoes)),
        ('Data da varredura', datetime.now().strftime('%d/%m/%Y %H:%M')),
    ]

    for i, (label, valor) in enumerate(dados_resumo, 3):
        ws_res.cell(row=i, column=1, value=label).font = Font(name='Arial', bold=True, size=11)
        ws_res.cell(row=i, column=2, value=valor).font = Font(name='Arial', size=11)
        ws_res.row_dimensions[i].height = 22

    ws_res.column_dimensions['A'].width = 25
    ws_res.column_dimensions['B'].width = 25

    wb.save(caminho_xlsx)
    print(f"\n  {c('Excel gerado:', Cor.GREEN + Cor.BOLD)} {caminho_xlsx}")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    # Argumentos
    if len(sys.argv) < 2:
        print(f"\n{c('Uso:', Cor.BOLD)} python scanner_rede.py <PASTA> [--top N] [--export]")
        print(f"\n{c('Exemplos:', Cor.DIM)}")
        print(f'  python scanner_rede.py "Z:\\Tarifas"')
        print(f'  python scanner_rede.py "Z:\\Tarifas" --top 20 --export')
        print(f'  python scanner_rede.py "\\\\servidor\\compartilhamento\\Tarifas"')
        sys.exit(1)

    pasta = sys.argv[1]
    top_n = TOP_PESADOS
    exportar = False

    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == '--top' and i + 1 < len(sys.argv):
            top_n = int(sys.argv[i + 1])
        if arg == '--export':
            exportar = True

    if not os.path.isdir(pasta):
        print(f"\n{c('ERRO:', Cor.RED)} Pasta não encontrada: {pasta}")
        sys.exit(1)

    # Execução
    print(c('\n╔══════════════════════════════════════════════════════════╗', Cor.CYAN))
    print(c('║     SCANNER DE REDE — ARPE / TARIFAS                   ║', Cor.CYAN))
    print(c('╚══════════════════════════════════════════════════════════╝', Cor.CYAN))

    arquivos, erros = escanear_pasta(pasta)

    if not arquivos:
        print(f"  {c('Nenhum arquivo encontrado na pasta.', Cor.YELLOW)}")
        sys.exit(0)

    duplicados = encontrar_duplicados_hash(arquivos)
    versoes = encontrar_versoes(arquivos)
    pesados = top_pesados(arquivos, top_n)

    exibir_duplicados(duplicados)
    exibir_versoes(versoes)
    exibir_pesados(pesados)
    exibir_resumo(arquivos, duplicados, versoes, erros)

    if exportar:
        exportar_excel(duplicados, versoes, pesados, arquivos, pasta)

    if erros:
        print(f"  {c('Arquivos inacessíveis (permissão negada):', Cor.DIM)}")
        for caminho, erro in erros[:5]:
            print(f"    • {caminho}")
        if len(erros) > 5:
            print(f"    ... e mais {len(erros) - 5}")
        print()


if __name__ == '__main__':
    main()